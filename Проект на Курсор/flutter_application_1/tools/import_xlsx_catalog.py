#!/usr/bin/env python3
"""
Читает Excel с колонкой наименований (как «Загрузка названий .xlsx») и генерирует
assets/catalog/imported_products.json для первого запуска приложения.

Зависимость: pip install openpyxl

Пример:
  python3 tools/import_xlsx_catalog.py "/path/to/file.xlsx"
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("Установите openpyxl: pip install openpyxl") from None

# Соответствует seedCategories() в mock_catalog_seed.dart
def infer_category_and_brand(name: str) -> tuple[str, str]:
    n = name.lower().strip()

    if re.match(r"^\d{2}\s+\d{3}\s+", n):
        return "iphone", "apple"
    if "iphone" in n:
        return "iphone", "apple"
    if "ipad" in n:
        return "ipad", "apple"
    if "macbook" in n:
        return "macbook", "apple"
    if "airpods" in n:
        return "airpods", "apple"
    if (
        "apple watch" in n
        or n.startswith("aw ")
        or " se2 " in n
        or n.startswith("se2 ")
        or " se 3 " in n
        or n.startswith("se 3 ")
        or "ultra " in n
        or re.search(r"\d{2,3}mm ", n)
        or " sb " in n
    ):
        return "apple_watch", "apple"
    if "dyson" in n or re.match(r"^hs\d", n) or re.match(r"^v\d", n) or " detect" in n:
        return "dyson", "dyson"
    if "samsung" in n or "galaxy" in n:
        return "samsung", "samsung"
    if "sony" in n or "playstation" in n or "ps5" in n or "wh-1000" in n:
        return "sony", "sony"
    if "instax" in n or "fujifilm" in n:
        return "instax", "fujifilm"
    if "steam deck" in n:
        return "steam_deck_oled", "valve"
    if "nintendo" in n or "switch" in n:
        return "nintendo_switch_2", "nintendo"
    if "meta quest" in n or "oculus" in n:
        return "meta_quest", "meta"
    if "bose" in n:
        return "bose", "bose"
    if "dji" in n:
        return "dji", "dji"
    if "garmin" in n:
        return "garmin", "garmin"
    if "gopro" in n:
        return "gopro", "gopro"
    if "marshall" in n:
        return "marshall", "marshall"

    return "accessories", "apple"


def is_section_or_junk(s: str) -> bool:
    t = s.strip()
    if not t or t == "Наименование":
        return True
    if re.fullmatch(r"\d+", t):
        return True
    if isinstance(s, (int, float)) and not isinstance(s, bool):
        return True
    low = {
        "iphone",
        "ipad",
        "accs",
        "dyson",
        "наименование",
        "macbook/ipad",
        "samsung",
        "sony",
        "bose",
        "meta quest",
        "nintendo",
        "steam deck",
        "instax",
        "garmin",
        "gopro",
        "dji",
        "marshall",
    }
    if t.lower() in low:
        return True
    if t.startswith(("⌚", "🎧", "💻")):
        return True
    if re.match(r"^Apple Watch", t) and "mm" not in t and "SB" not in t and "GB" not in t:
        return True
    # Подзаголовки линеек без конкретной комплектации
    if re.match(r"^iPad Air \d+ M\d+$", t):
        return True
    if re.match(r"^iPad Pro", t) and "wi-fi" not in t.lower() and "cellular" not in t.lower():
        return True
    return False


def picsum_seed(name: str) -> str:
    h = hashlib.md5(name.encode("utf-8")).hexdigest()[:16]
    return f"https://picsum.photos/seed/{h}/900/1100"


def extract_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = [r[0] for r in wb.active.iter_rows(values_only=True)]
    wb.close()
    out: list[str] = []
    for cell in rows:
        if cell is None:
            continue
        if isinstance(cell, (int, float)) and not isinstance(cell, bool):
            if cell == int(cell) and 1 <= cell <= 500:
                continue
        s = str(cell).strip()
        if is_section_or_junk(s):
            continue
        out.append(s)
    return out


def build_products(names: list[str]) -> list[dict]:
    products: list[dict] = []
    for i, name in enumerate(names):
        cat, brand = infer_category_and_brand(name)
        pid = f"imp_{i + 1}_{hashlib.md5(name.encode()).hexdigest()[:10]}"
        products.append(
            {
                "id": pid,
                "name": name,
                "categoryId": cat,
                "brandId": brand,
                "price": 0.0,
                "oldPrice": None,
                "currency": "RUB",
                "images": [picsum_seed(name)],
                "description": "Позиция загружена из прайса. Укажите цену и наличие в админке.",
                "specifications": {},
                "specificationsMap": {"Источник": "Импорт XLSX"},
                "stockQuantity": 1,
                "rating": 0.0,
                "isPopular": False,
                "isNew": False,
                "variants": [],
                "sku": f"IMP-{i + 1:05d}",
            }
        )
    return products


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "xlsx",
        nargs="?",
        default=str(
            Path.home()
            / "Downloads"
            / "Загрузка названий .xlsx"
        ),
        help="Путь к .xlsx",
    )
    args = ap.parse_args()
    src = Path(args.xlsx).expanduser().resolve()
    if not src.is_file():
        raise SystemExit(f"Файл не найден: {src}")

    root = Path(__file__).resolve().parents[1]
    out_path = root / "assets" / "catalog" / "imported_products.json"

    names = extract_names(src)
    products = build_products(names)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(products, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"OK: {len(names)} позиций → {out_path}")


if __name__ == "__main__":
    main()
